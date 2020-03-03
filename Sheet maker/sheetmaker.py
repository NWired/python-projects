import openpyxl,os,re,pathlib

#openpyxl: to handle excel sheet files

#.xlsx files are excel sheet files
#import re: regex needed to check formats (to identify which are xlsx files)
#pathlib: for traversing through the current directory
from openpyxl.styles import Font #for the title
from pathlib import Path #especially to get current directory
from prettytable import PrettyTable #this is to display the sheet contents
import pyinputplus as pyip #for inputs

def create_file():
    print('This is an example:')
    view_sheet('sample.xlsx')
    creating = pyip.inputYesNo('Adding?[Y/N]')
    
    if creating == 'yes':
        filename = input('File name:') #name of file
        #for the titles of each heading
        fontsetting = Font(bold = True) 
        titles = int(input('Number of titles:')) 

        #no of sets of items 
        items = int(input('Number of items:'))
        wb = openpyxl.Workbook() 
        current = wb.active #current sheet
        #row is vertical, column is horizontal
        
        for i in range(titles): #add titles
            title = input('title:')
            current.cell(row = 1,column = i+1).font = fontsetting
            current.cell(row = 1,column = i+1).value = title
            
        for j in range(items): #to add items
            for k in range(titles): 
                item = input('item:')
                current.cell(row = j+2,column = k+1).value = item
            
        wb.save('%s.xlsx'%filename)
    
def edit_sheet(chosen_file): #edit sheets
    option = 'yes'
    while option == 'yes': #choose what sheet
        openbook = openpyxl.load_workbook(chosen_file)
        sheets = openbook.sheetnames
        chosen_sheet = None
    
        if len(sheets) >= 2:
            chosen_sheet = pyip.inputMenu(sheets,numbered = True)
        elif len(sheets) == 1:
            chosen_sheet = sheets[0]
        else:
            print('N/A')
        if chosen_sheet != None:
            chosen_sheet_obj = openbook[chosen_sheet] #change it into sheet obj
            view_sheet(chosen_file,chosen_sheet)
            chosen_row = pyip.inputInt('Choose a row:')
            chosen_column = pyip.inputInt('Choose a column:')
            new_value = pyip.inputStr('New value:')

            chosen_sheet_obj.cell(\
                row = chosen_row,column = chosen_column ).value = new_value
            
            openbook.save(chosen_file)
            option = pyip.inputYesNo('Anymore to edit?')
            
def choose_file(): #return sheet name
    files = []
    actions = ['view_sheet','add_sheet','edit_sheet','delete_sheet','back'] #option
    current = Path.cwd()
    fileregex = re.compile('([A-Za-z1-9\-*_])+.xlsx')
    
    for folderName, subfolders, filenames in os.walk(current):
        for filename in filenames:
            if fileregex.match(filename) != None:
                files.append(filename)
                
    action = pyip.inputMenu(actions,numbered = True) #choose action

    
    if action == 'back': #go back
        eval('%s()'%str(action))
    else:
        filename = pyip.inputMenu(files,numbered = True) #chosen file
        eval('%s("%s")'%(action,str(filename)))

#add sheet
def add_sheet(chosen_file):
    print('Add sheets')
    option = 'yes'
    while option == 'yes': #choose what sheet
        openbook = openpyxl.load_workbook(chosen_file) #the main workbook
        sheets = openbook.sheetnames
        add = pyip.inputYesNo('To add?[Y/N]')       
        if add == 'yes':
            sheet_name = pyip.inputStr('New sheet name:')
            openbook.create_sheet(sheet_name)
            openbook.save(chosen_file)
            print('%s is add.'%sheet_name)
        openbook.save(chosen_file)
        
        option = pyip.inputYesNo('Anymore to added?')

        


#delete sheet
def delete_sheet(chosen_file):
    print('Delete sheets')
    option = 'yes'
    while option == 'yes': #choose what sheet
        openbook = openpyxl.load_workbook(chosen_file) #the main workbook
        sheets = openbook.sheetnames
        chosen_sheet = None
    
        if len(sheets) >= 2:
            chosen_sheet = pyip.inputMenu(sheets,numbered = True)
        elif len(sheets) == 1:
            chosen_sheet = sheets[0]
        else:
            print('N/A')
        if chosen_sheet != None:
            chosen_sheet_obj = openbook[chosen_sheet] #change it into sheet obj
            view_sheet(chosen_file,chosen_sheet)
        delete = pyip.inputYesNo('To delete?[Y/N]')
        
        if delete == 'yes':            
            del openbook[chosen_sheet]
            openbook.save(chosen_file)
            print('%s is deleted.'%chosen_sheet)
        
        option = pyip.inputYesNo('Anymore to delete?')
    
#choose sheet
def view_sheet(workbook,chosen_sheet = None):
    openbook = openpyxl.load_workbook(workbook)
    sheets = openbook.sheetnames
    if chosen_sheet == None:
    ##    option = inputYesNo() #still add 
        if len(sheets) >= 2:
            chosen_sheet = pyip.inputMenu(sheets,numbered = True)
        elif len(sheets) == 1:
            chosen_sheet = sheets[0]
        else:
            print('N/A')

 
    #row is vertical, column is horizontale
    items = [ ]
    rows = openbook[chosen_sheet].max_column #horizontal
    columns = openbook[chosen_sheet].max_row   #vertical

    for i in range(1,columns+1): #y
        row_items = []
        row_items.append(str(i))
        for j in range(1,rows+1): #x
            row_items.append(openbook[chosen_sheet].cell(column = j,\
                                                     row = i).value)
        items.append(row_items)
    #to display the items there
    table = PrettyTable()
    table.field_names = ['Row/Column'] + [ a for a in range(1,rows+1)]
    for i in items:
        table.add_row(i)

##    sort = pyip.inputYesNo('Sort?[Y/N]')
##    if sort == 'yes':
##        category = 
##        table.sortby = 
    print('File name:%s\nSheet name:%s'%(workbook,chosen_sheet))
    print(table)

            

    

def back():
    options = ['create_file','choose_file','exit']
    item = ''
    while item != 'exit':
        item = pyip.inputMenu(options,numbered = True)
        eval('%s()'%str(item))

back()

#welcome
    #create xlsx file 
    #choose xlsxfiles
        #create sheet (done)
        #edit sheet (done)
    #read sheets (done)
    
